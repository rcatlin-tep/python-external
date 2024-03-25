# -*- coding: utf-8 -*-
"""
Created on Mon Aug 01 2022
@author: AA60857
"""

from pi.core.settings import MultiDomainPiSettings
from pi.core.PiConnectionManager import PiConnectionManager
import pandas as pd

# import sys 
# import clr
# sys.path.append(r'C:\Program Files (x86)\PIPC\AF\PublicAssemblies\4.0')  
# clr.AddReference('OSIsoft.AFSDK')  
# from OSIsoft.AF import *  
# from OSIsoft.AF.PI import *  
# from OSIsoft.AF.Asset import *  
# from OSIsoft.AF.Data import *  
# from OSIsoft.AF.Time import *  
# from OSIsoft.AF.UnitsOfMeasure import *
# import pytest
# import openpyxl
# sys.path.append('../../')  
# from tepcorepidata import tepcorepidata
# from teppwcore import teppwcore
# from tepcore import tepcore

from tepcorepidata import tepcorepidata
from tepcore import tepcore
from datetime import datetime
from teppwcore import teppwcore
import sys 
import clr
sys.path.append(r'C:\Program Files (x86)\PIPC\AF\PublicAssemblies\4.0')    
clr.AddReference('OSIsoft.AFSDK')  
from OSIsoft.AF import *  
from OSIsoft.AF.PI import *  
from OSIsoft.AF.Asset import *  
from OSIsoft.AF.Data import *  
from OSIsoft.AF.Time import *  
from OSIsoft.AF.UnitsOfMeasure import *
import pytest
import openpyxl
# =============================================================================
# import clr
# import sys
# clr.AddReference("System.Collections")
# from System.Collections.Generic import Dictionary
# from System import String
# from System import Object
# sys.path.append(r'C:\Program Files (x86)\PIPC\AF\PublicAssemblies\4.0')    
# clr.AddReference('OSIsoft.AFSDK')  
# from OSIsoft.AF import *  
# from OSIsoft.AF.PI import *  
# from OSIsoft.AF.Asset import *  
# from OSIsoft.AF.Data import *  
# from OSIsoft.AF.Time import *  
# from OSIsoft.AF.UnitsOfMeasure import *
# from OSIsoft.AF.Search import *
# =============================================================================

class EmsCorpTagManager:
    #standard EMS-CORP Feilds
    location1 = '1'
    location2 = '0'
    location3 = '7'
    location4 = '1'
    location5 = '0'
    pointsource = 'EMS-CP-RT'
    ptclassname = 'classic'
    
    def __init__(self,debug ,domain='dev',osissystem='ems',config='piconnectionmanager.json',usepi=True):
        '''
        A class for analyzing and creating tags across PI systems. Designed for TEP CORP and TEP EMSSEC pi system integrations

        Parameters
        ----------
        debug : TYPE
            DESCRIPTION.
        domain : TYPE, optional
            DESCRIPTION. The default is 'dev'.
        osissystem : TYPE, optional
            DESCRIPTION. The default is 'ems'.
        config : TYPE, optional
            DESCRIPTION. The default is 'piconnectionmanager.json'.
        usepi : TYPE, optional
            DESCRIPTION. The default is True.

        Returns
        -------
        None.

        '''
        self.warnings = []
        
        self.osissystem = osissystem
        self.domain = domain
        self.pcm = PiConnectionManager(debug,domain,osissystem,config,usepi)
        self.pidatacorp = self.pcm.GetCorpPiConnection()
        self.pidataems = self.pcm.GetEmsPiConnection()    
        self.settings = self.pcm.settings
        self.tc = self.pcm.tc
        
    def GetEMSPointListWithAttributesByDate(self,generic=10):
        daysback = int(generic) # Days back to search based in the creation date
        points = []
        tag = None
        descriptior=None
        pointtyupe=None
        digitalset=None
        typicalvalue = 0.0
        instrumenttag = None
        pointsouce = None
        isquality = 0

        if self.pidataems is not None:
            tagsEMSPI = self.pidataems.GetPointListByAttribute('creationdate','',daysback,0,True) #attribute to search, string to search, daysback to search in creation date, exact match search
            #tagEMSPI.LoadAttributes('tag','descriptor','pointtype','digitalset','typicalvalue','creationdate')
            for tagEMSPI in tagsEMSPI:
                tagandattributes = self.pidataems.GetPointStandardAttributes(tagEMSPI)
                for attribute in tagandattributes:
                    if attribute[0] == 'tag':
                        tag = attribute[1] 
                        if '@DATA_QUALITY' in tag:
                            isquality = 1
                        else:
                            isquality = 0
                    elif attribute[0] == 'descriptor':
                        descriptor = attribute[1] 
                    elif attribute[0] == 'pointtype':
                        pointtype = attribute[1] 
                    elif attribute[0] == 'digitalset':
                        digitalset = attribute[1] 
                    elif attribute[0] == 'typicalvalue':
                        typicalvalue = attribute[1] 
                    elif attribute[0] == 'instrumenttag':
                        instrumenttag = attribute[1] 
                    elif attribute[0] == 'pointsource':
                        pointsource = attribute[1] 
                points.append([tagEMSPI,tag,descriptor,pointtype,digitalset,typicalvalue,instrumenttag,pointsource,isquality])

            return pd.DataFrame(points,columns=['Name','tag','descriptor','pointtype','digitalsetname','typicalvalue','instrumenttag','pointsource', 'isquality'])
        return None
    
    def GetEMSPointsListWithAttributes(self):
        res = self.pidataems.GetAllPointListAsDFWithStandardAttributes('')
        return res
    
    
    def GetCorpPointsListWithAttributes(self):
        res = self.pidatacorp.GetAllPointListAsDFWithStandardAttributes('')
        return res
    
    def ApplyFiltersForEMSDataframe(self,inframe):
            '''
            
    
            Parameters
            ----------
            inframe : TYPE
                A data frame of points to filter.
    
            Returns
            -------
            TYPE
                Filtered data frame of HSH-DIST and HSH-EMS , no @DATA_QUALITY and NO :5min.
    
            '''
            #inframe.query('''pointsource == 'HSH-DIST' and isquality==0 ''')
            
            return inframe.query('''pointsource in ('HSH-DIST','HSH-EMS') and isquality==0 and ~tag.str.contains(":5MIN")  ''')
    
    
    
    
    def FindPointsNotINCorp(self,daysback):

        self.tc.debug('Retrieving tags from EMS PI...')
        tagsdf = self.GetEMSPointListWithAttributesByDate(daysback)
        if self.tc.isValidDf(tagsdf):
            tagsdf = self.ApplyFiltersForEMSDataframe(tagsdf)
            if self.tc.isValidDf(tagsdf):
                self.tc.debug('We have new EMS tags to evalualte agisnt corporate pi')
                print(tagsdf.head())
                #get the corp point list:
                     
                corppointlistdf = self.GetCorpPointsListWithAttributes()
                if self.tc.isValidDf(corppointlistdf):
                     self.tc.debug('Succesfully loaded all corp tags for comparison')
                     dfout=tagsdf.merge(corppointlistdf, how='left',indicator=True,left_on=['tag'], right_on=['instrumenttag'])
                     new_points_df = dfout.query('''_merge=='left_only' ''')
                     if self.tc.isValidDf(new_points_df):
                         self.tc.debug('We have new tags to create in CORP PI')
                         '''Name_x                object
descriptor_x          object
tag_x                 object
pointtype_x            int64
digitalsetname_x      object
typicalvalue_x       float64
instrumenttag_x       object
pointsource_x         object'''
                         #descriptor, tag, pointtype, location1, location2, location3, location4, location5, digitalset, pointsource, ptclassname, typicalvalue
                         #columns=['descriptor','tag','pointtype','location1', 'location2', 'location3', 'location4', 'location5', 'digitalset', 'pointsource', 'ptclassname', 'typicalvalue']
                         new_points_df = new_points_df[['descriptor_x','tag_x','pointtype_x','digitalsetname_x','typicalvalue_x']]
                         new_points_df.rename(columns={'descriptor_x':'descriptor','tag_x':'tag','pointtype_x':'pointtype','digitalsetname_x':'digitalset','typicalvalue_x':'typicalvalue'},inplace=True)


                         new_points_df['location1'] = self.location1
                         new_points_df['location2'] = self.location2
                         new_points_df['location3'] = self.location3
                         new_points_df['location4'] = self.location4
                         new_points_df['location5'] = self.location5
                         new_points_df['pointsource'] = self.pointsource
                         new_points_df['ptclassname'] = self.ptclassname
                         self.tc.debug(new_points_df.head())
                         return [True,new_points_df]
                     else:
                         self.tc.debug('All points match')
                         return [True,None]
                 
                else:
                     self.tc.debug('Unable to find copr point list exiting fattaly')
                     return [False,None]
            else:
                self.tc.debug('No new tags after filter process')
                return [True,None]
        else:
            self.tc.debug('No new tags to process')
            return [True,None]
    
    
    
    def EmsPiToCorp(self,generic=10):
        daysback = int(generic) # Days back to search based in the creation date

    
        self.tc.debug('Starting looking the tags in EMS PI based in the creation date' + str(daysback) + 'day(s) back')
        
        res = self.FindPointsNotINCorp(daysback)
        
        if res[0]:
            if self.tc.isValidDf(res[1]):
                self.tc.debug('Loading points')
                self.DoCreatePIPointsCorp(res[1])
            return True
        else:
            return False
    
        
                        
    def DoCreatePIPointsCorp(self, dataframe):

        return self.pidatacorp.CreatePIPointsWithAttibutesFromDataframe(dataframe)
    
    
    def FindDuplicateTags(self):
        from openpyxl import Workbook
        
        wb = Workbook()
        points_data = self.GetCorpPointsListWithAttributes()
        indices = points_data[((points_data['instrumenttag'].str.contains(':D:') == False) & (points_data['instrumenttag'].str.contains(':E:') == False)) | \
                              (points_data['instrumenttag'].str.contains('@') == True)].index
            
        points_data.drop(indices, inplace=True)
        points_data['Key'] = points_data['instrumenttag'].str[-8:]
        points_data['PureKey'] = points_data['instrumenttag'].str[-10:]
        print(points_data.count)
        pure_duplicates = points_data[points_data.duplicated('PureKey') | points_data.duplicated('PureKey', keep='last')]
        weak_duplicates = points_data[points_data.duplicated('Key') | points_data.duplicated('Key', keep='last')]
        weak_duplicates = pd.merge(weak_duplicates, pure_duplicates, indicator=True, how='outer').query('_merge=="left_only"').drop('_merge', axis=1)
        
        ws = wb.active;
        ws.append(['Sort with this!', 'name', 'descriptor', 'instrumenttag', 'Flag exact match', 'Flag key duplicate'])
        for index, row in pure_duplicates.iterrows():
            if row['instrumenttag'] is not None:
                ws.append([row['Key'], row['Name'], row['descriptor'], row['instrumenttag'], 1, 0])
        for index, row in weak_duplicates.iterrows():
            if row['instrumenttag'] is not None:
                ws.append([row['Key'], row['Name'], row['descriptor'], row['instrumenttag'], 0, 1])
        wb.save('./Reports/DuplicateInstrumentTags/duplicate_tags.xlsx')
        
    def FindTagsInAF(self):
        piafserver = 'tuswppiaf01'
        piserver = 'TUSWPPI01'
        piafdb = 'TEPPIUDS'
        piafdata = tepcorepidata(piafserver,True,True,piafdb,True,piserver)
    
        dbinfo = teppwcore('qa') 
        res = piafdata.GetAfToTagMapping('*', False)
        if res is not None:
            wb = openpyxl.load_workbook('./Reports/DuplicateInstrumentTags/duplicate_tags.xlsx')
            ws = wb.active
            cur_name = 'nullnullnullnull'
            cur_row = 1;
            while cur_name is not None:
                result = res[res['afconfigstring'].str.contains(cur_name)]
                if not result.empty:
                    print(result)
                    ws.cell(row=cur_row, column=8).value = '1'
                cur_row += 1
                cur_name = ws.cell(row=cur_row, column=2).value
            wb.save('./Reports/DuplicateInstrumentTags/found_tags.xlsx')

